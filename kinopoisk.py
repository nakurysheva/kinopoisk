

def namecells():# The fuction to name cells in the Excel sheet
    ws['A1'] = "Movie Name"
    ws['B1'] = "Eng movie name"
    ws['C1'] = "Year"
    ws['D1'] = "Country"
    ws['E1'] = "Running time in min."
    ws['F1'] = "Cast"
    ws['G1'] = "Director"
    ws['H1'] = "Genre"


def stringOnly(htmlstr, catname):#The function picks strings from a piece of html tree and returns this string under necessary category name
    for i in htmlstr:
        catname = (str(i))
    return catname


def movieInfoKinoPoisk(data):# The function to parse necessary information from html tree
    moviename = data.p.a# Parsing name of the movie which is contained within tag 'a' under tag 'p'
    MovieName = ""
    MovieName = stringOnly(moviename, MovieName)
    year = data.find('span', class_="year")# Parsing year which is contained within tag 'span' class="year"
    Year = ''
    Year = stringOnly(year, Year)
    enMoviename = data.find('span', class_="gray")# Parsing English name of the movie which is contained within tag 'span' class="gray"
    for string in enMoviename:
        string =str(string)
# Due to the fact that within tag 'span' class="gray" there were two units of necessary information (English name and duration of the movie)
# the string should be split into two parts. In order to check if both units of information are there in the string and to save them in respective cells of Ecxel
# the following block of 'if' statements was created
        if string.find(",") != -1:# If there is a comma in the string
            if string.find(",") == string.rfind(","):# check if there is only one comma
                string = string.split(',')# split the string by the comma
                EnMovieName = (string[0])
                time = string[1].strip()
                minutes = time.split()# extracting only numbers from the duration segment
                Minutes = (minutes[0])
            elif string.find("мин") != 1:# checking if there is only duration unit
                index = string.rfind(',')
                EnMovieName = string[:index]
                time = string[index + 1:]
                time = time.strip()
                minutes = time.split()
                Minutes = (minutes[0])
            else:
                EnMovieName = str(string)#if all the above is not True, there is only name of the movie without duration
                Minutes = str(None)
        elif string.find("мин") != -1:# check if there is duration unit
            EnMovieName = MovieName
            time = string.strip()
            minutes = time.split()
            Minutes = (minutes[0])
        else:
            EnMovieName = str(string)
            Minutes = str(None)
    country = enMoviename.find_next('span', class_="gray")# Parsing Country which is contained within the second tag 'span' class="gray" in the html tree
    two_units = []#in this tag there are two units of necessary information
    for i in country:
        two_units.append(str(i))
    Genre = two_units[4]# Extracting genre placed on the 5th position in our list
    list_symb = [' ', ')', '(', '...']# and deleting unnecessary symbols from genre
    for i in list_symb:
        Genre = Genre.replace(i, ' ')
        Genre = Genre.strip()
    NameCountry = str(two_units[0])# Extracting country name placed on the 1st position in our list
    NameCountry = NameCountry[:-2]
    director = country.find('a', class_="lined js-serp-metrika")# Parsing director name from tag 'a' class="lined js-serp-metrika"
    Director = ''
    Director = stringOnly(director, Director)
    starring = director.find_next('a', class_="lined js-serp-metrika")# Parsing cast from the next tag 'a' class="lined js-serp-metrika"
    list_star = []# As there are two names of main actors under two similar tags, a list of stars is needed
    Star1 = ''
    Star1 = stringOnly(starring, Star1)
    list_star.append(Star1)# Adding 1st actor to the list
    starring2 = starring.find_next('a', class_="lined js-serp-metrika")# Parsing 2nd actor from the next tag 'a' class="lined js-serp-metrika"
    Star2 = ''
    Star2 = stringOnly(starring2, Star2)
    list_star.append(Star2)# Adding 2nd actor to the list
    stars = ''
    for i in list_star:
        a = str(i)
        stars += a + ", "
    stars = stars[:-2]
    return MovieName, Year, EnMovieName, Minutes, Genre, NameCountry, Director, stars# The function returns 8 units of information


from bs4 import BeautifulSoup
from urllib.request import urlopen
from openpyxl import Workbook


#Creating new workbook, open first sheet and name cells with the help of the function 'namecells'.
wb = Workbook()
ws = wb.get_sheet_by_name("Sheet")
namecells()

#Recieving html tree from web page with the help of BeautifulSoup library
html_doc = urlopen("https://www.kinopoisk.ru/s/type/film/list/1/m_act%5Bgenre%5D%5B0%5D/18/m_act%5Bgenre_and%5D/on/").read()
soup = BeautifulSoup(html_doc, "html5lib")
info = soup.find_all('div', 'info')# Selecting part of html tree where necessary information is stored

#Cycle initiating search in the html tree and placing found information into respective cells of the Excel file
count = 2
for i in info:
    ws['A' + str(count)], ws['C' + str(count)], ws['B' + str(count)], ws['E' + str(count)], ws['H' + str(count)], ws['D' + str(count)], ws['G' + str(count)], ws['F' + str(count)] = movieInfoKinoPoisk(i)
    count += 1
wb.save('KinoPoisk.xlsx')  #Saving results to the Excel file
